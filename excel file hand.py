#write data
"""from openpyxl  import workbook
wb=workbook()
#select active sheet
sheet=wb.active
#add data in sheet
sheet["A1"]="Name"
sheet["B1"]="Age"
sheet["C1"]="Marks"

sheet["A2"]="Ali"
sheet["B2"]=20
sheet["C2"]=85
#save file
wb.save("student.xlsx")"""
#read data
from openpyxl import load_workbook
wb=load_workbook("ex.xlsx")
sheet=wb.active
"""print(sheet["A1"].value)
print(sheet["A2"].value)
print(sheet["B1"].value)
print(sheet["B2"].value)

print(sheet["C1"].value)
print(sheet["C2"].value)"""
for row in sheet.iter_rows(values_only=True):
    print(row)
