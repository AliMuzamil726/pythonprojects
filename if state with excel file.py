from openpyxl import Workbook, load_workbook

print("Select Mode:")
print("1. Write Mode")
print("2. Read Mode")
print("3. Append Mode")

choice = input("Enter your choice (1/2/3): ")

# ---------------- WRITE MODE ----------------
if choice == "1":
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Ali"
    sheet["B1"] = 18
    sheet["C1"] = 20
    wb.save("ali.xlsx")
    wb.close()
    print("Write mode completed")

# ---------------- READ MODE ----------------
elif choice == "2":
    wb = load_workbook("students.xlsx")
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
    wb.close()

# ---------------- APPEND MODE ----------------
elif choice == "3":
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Name", "Age", "Marks"])
    sheet.append(["Ali", 19, 80])
    sheet.append(["Muzamil", 20, 78])
    sheet.append(["usman", 57, 66])
    sheet.append(["iqbal", 28, 90])

    wb.save("ali.xlsx")
    wb.close()
    print("Append mode completed")

# ---------------- INVALID INPUT ----------------
else:
    print("Invalid choice")
