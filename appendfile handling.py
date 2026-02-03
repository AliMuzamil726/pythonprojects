from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

sheet.append(["Name", "Age", "Marks"])
sheet.append(["Usman", 20, 85])
sheet.append(["Ali", 21, 90])
sheet.append(["Muzamil",25,78])

wb.save("students.xlsx")
wb.close()


from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
note = wb.active
for row in note.iter_rows(values_only=True):
    print(row)
wb.close()
#for find number of rows and columns

print(sheet.max_column)
print(sheet.max_row)
#create new sheet in excel file

sheet2=wb.create_sheet("Alii.xlsx")
wb.save("students.xlsx")
#for specific column  
for cell in sheet["A"]:
 print(cell.value)
#for specific  row
for cell in sheet["1"]:
 print(cell.value)